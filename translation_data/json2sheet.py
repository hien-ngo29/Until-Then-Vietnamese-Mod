from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Border, Side
from pathlib import Path
import pandas, json, sys

NON_TRANSLATABLE_EXACT = {
    "::",
    ":: ",
    "1",
    "2",
    "3",
    "",
    " ",
}

NON_TRANSLATABLE_PREFIXES = (
    "NOP",
    "$",
    "type: ",
    "player: ",
    "characters: ",
    "id: ",
    "music_",
    "ids: ",
    "status_",
    "X",
    "hide_back_btn:",
    "push_focus:",
    "\""
)

def write_xlsx(jsons_dir: Path, output_dir: Path=Path("sheet")):
    excel_writer = pandas.ExcelWriter(output_dir / "UntilThen.xlsx", engine="openpyxl")
    
    workbook = excel_writer.book
    worksheet = workbook.create_sheet("UntilThen")
    
    write_headers_to_sheet(excel_writer, worksheet)

    prev_dataframe = None
    sum_row = 0
    write_row = 0

    for i in range(0, len(all_filenames)):
        filename = all_filenames[i]
        dict_data = get_dict_data_from_json_file(filename)

        dataframe = pandas.DataFrame([dict_data]).T
        write_row = (write_row + len(prev_dataframe)) if (prev_dataframe is not None) else 0
        
        sum_row += len(dict_data)

        dataframe.to_excel(excel_writer, sheet_name="UntilThen", startrow=write_row+1, header=False)
        
        non_translatable_quality_dataframe = pandas.DataFrame([{"quality": -1}])
        
        for i, sentence in enumerate(dict_data.keys()):
            if (not is_translatable_sentence(sentence)):
                non_translatable_quality_dataframe.to_excel(excel_writer, sheet_name="UntilThen", startrow=write_row+1+i, startcol=2, header=False, index=False)

        prev_dataframe = dataframe
        
        update_loading_bar(float(sum_row * 100 / 45284))
    
    set_xlsx_color_format(excel_writer)
    apply_all_borders(worksheet)
    
    print("Done! Dumped " + str(sum_row) + " lines")

    excel_writer.close()
    
def write_headers_to_sheet(excel_writer, worksheet):
    excel_writer.sheets["UntilThen"].column_dimensions["A"].width = 90
    excel_writer.sheets["UntilThen"].column_dimensions["B"].width = 90
    excel_writer.sheets["UntilThen"].column_dimensions["C"].width = 15
    excel_writer.sheets["UntilThen"].column_dimensions["D"].width = 90

    worksheet["A1"] = "VĂN BẢN GỐC"
    worksheet["B1"] = "VĂN BẢN DỊCH"
    worksheet["C1"] = "CHẤT LƯỢNG"
    worksheet["D1"] = "NOTE"

def apply_all_borders(worksheet):
    thin = Side(style="thin")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=worksheet.max_column
    ):
        for cell in row:
            cell.border = border
    
def is_translatable_sentence(string: str):
    if (string.startswith("$ thought") and string != "$ thought \""):
        return True
    
    if (string[0].islower() and len(string.split(" ")) == 1):
        return False
    
    if string in NON_TRANSLATABLE_EXACT:
        return False

    if string.startswith(NON_TRANSLATABLE_PREFIXES):
        return False

    return True

def update_loading_bar(percent, width=30):
    percent = max(0, min(percent, 100))
    filled = int(width * percent / 100)

    bar = "#" * filled + "-" * (width - filled)

    sys.stdout.write(f"\r[{bar}] {percent:.1f}%")
    sys.stdout.flush()

    if percent >= 100:
        print()
        
def set_xlsx_color_format(excel_writer):
    worksheet = excel_writer.sheets["UntilThen"]

    grey = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    column_range = "C2:C45284"

    worksheet.conditional_formatting.add(
        column_range,
        CellIsRule(operator="equal", formula=['-1'], fill=grey)
    )

    worksheet.conditional_formatting.add(
        column_range,
        CellIsRule(operator="equal", formula=['1'], fill=green)
    )

    worksheet.conditional_formatting.add(
        column_range,
        CellIsRule(operator="equal", formula=['2'], fill=yellow)
    )

    worksheet.conditional_formatting.add(
        column_range,
        CellIsRule(operator="equal", formula=['3'], fill=red)
    )

all_filenames = []
file = open("test.txt", "w")
def get_all_json_filenames(jsons_dir: Path):
    for filepath in jsons_dir.iterdir():
        if filepath.is_file():
            all_filenames.append(filepath)
        else:
            get_all_json_filenames(filepath)

def get_dict_data_from_json_file(json_path: Path):
    file = open(str(json_path), "r")
    return json.load(file)

get_all_json_filenames(Path("json"))
write_xlsx(jsons_dir=Path("json"))
file.close()